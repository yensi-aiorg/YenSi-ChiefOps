"""
Per-project file upload processing.

Handles files uploaded to a specific project: Slack JSON exports, Jira XLSX
spreadsheets, and general documentation (PDF, DOCX, MD, TXT). Each file
produces a ``text_documents`` entry tagged with ``project_id`` and is
optionally ingested into Citex for RAG retrieval.
"""

from __future__ import annotations

import io
import json
import logging
import tempfile
from pathlib import Path
from typing import TYPE_CHECKING, Any

from app.citex.client import CitexClient
from app.citex.project_scope import derive_citex_project_id
from app.config import get_settings
from app.models.base import generate_uuid, utc_now
from app.services.ingestion.drive import (
    _extract_docx_text,
    _extract_pdf_text,
    _extract_plain_text,
    _get_content_type,
)
from app.services.ingestion.hasher import compute_hash
from app.services.insights.semantic import extract_semantic_insights, generate_project_snapshot

if TYPE_CHECKING:
    from motor.motor_asyncio import AsyncIOMotorDatabase

logger = logging.getLogger(__name__)

# Supported extensions for per-project upload
ALLOWED_EXTENSIONS: set[str] = {".json", ".xlsx", ".pdf", ".docx", ".md", ".txt"}

_TEXT_EXTENSIONS = {".txt", ".md", ".markdown"}


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


async def process_project_file(
    project_id: str,
    filename: str,
    content: bytes,
    db: AsyncIOMotorDatabase,  # type: ignore[type-arg]
) -> dict[str, Any]:
    """Process a single file uploaded to a project.

    Detects file category by extension, extracts text, stores metadata in
    ``project_files``, creates/upserts a ``text_documents`` entry with
    ``project_id``, and attempts Citex ingestion.

    Returns:
        A result dict with keys: file_id, filename, file_type, status,
        records_created, citex_ingested, error_message.
    """
    ext = Path(filename).suffix.lower()
    file_type = _detect_file_type(ext)
    content_hash = compute_hash(content)
    file_id = generate_uuid()
    now = utc_now()

    result: dict[str, Any] = {
        "file_id": file_id,
        "filename": filename,
        "file_type": file_type,
        "status": "completed",
        "records_created": 0,
        "citex_ingested": False,
        "error_message": None,
    }

    # Check duplicate within this project
    existing = await db.project_files.find_one(
        {"project_id": project_id, "content_hash": content_hash}
    )
    if existing:
        result["status"] = "skipped"
        result["error_message"] = "Duplicate file (same content already uploaded to this project)"
        result["file_id"] = existing["file_id"]
        return result

    try:
        # Dispatch to category-specific processor
        if file_type == "slack_json":
            text_content, records = await _process_slack_json(
                project_id, filename, content, db
            )
        elif file_type == "jira_xlsx":
            text_content, records = await _process_jira_xlsx(
                project_id, filename, content, db
            )
        else:
            text_content, records = await _process_documentation(
                project_id, filename, content, ext
            )

        result["records_created"] = records

        # Store text document if we extracted text
        text_document_id = None
        semantic_summary_text = ""
        if text_content and text_content.strip():
            text_document_id = await _store_text_document(
                project_id=project_id,
                source=file_type,
                source_ref=filename,
                title=filename,
                content=text_content,
                db=db,
            )
            semantic_result = await extract_semantic_insights(
                project_id=project_id,
                source_type=file_type,
                source_ref=filename,
                content=text_content,
                db=db,
            )
            semantic_summary_text = str(semantic_result.get("summary_text", "")).strip()
            await generate_project_snapshot(project_id=project_id, db=db, force=True)

        # Attempt Citex ingestion (skipped when CITEX_ENABLED=False)
        settings = get_settings()
        if settings.CITEX_ENABLED and text_content and text_content.strip():
            # Preserve raw source bytes for formats Citex parses by structure/containers.
            passthrough_upload_exts = {".json", ".xlsx", ".pdf", ".docx"}
            upload_bytes: bytes | None = content if ext in passthrough_upload_exts else None
            result["citex_ingested"] = await _ingest_to_citex(
                project_id=project_id,
                content=text_content,
                filename=filename,
                source=file_type,
                file_bytes=upload_bytes,
                content_type=_get_content_type(ext) if upload_bytes is not None else None,
            )
            if result["citex_ingested"] and text_document_id:
                await _record_citex_ingestion_state(
                    db=db,
                    project_id=project_id,
                    source_group="docs",
                    source=file_type,
                    source_ref=filename,
                    document_id=text_document_id,
                    content_hash=compute_hash(text_content.encode("utf-8")),
                )
                await db.text_documents.update_one(
                    {"document_id": text_document_id},
                    {
                        "$set": {
                            "indexed_in_citex": True,
                            "citex_project_id": derive_citex_project_id(
                                configured_project_id=settings.CITEX_PROJECT_ID,
                                api_key=settings.CITEX_API_KEY,
                                fallback_project_id=project_id,
                            ),
                            "citex_last_ingested_at": utc_now(),
                            "updated_at": utc_now(),
                        }
                    },
                )
            # Send semantic extraction output as a dedicated text document to Citex.
            if semantic_summary_text:
                summary_filename = f"{Path(filename).stem}__semantic_summary.txt"
                summary_content = (
                    f"Semantic summary for: {filename}\n\n"
                    f"{semantic_summary_text}\n"
                )
                summary_ok = await _ingest_to_citex(
                    project_id=project_id,
                    content=summary_content,
                    filename=summary_filename,
                    source="semantic_summary",
                    source_ref=filename,
                    content_type="text/plain",
                )
                if not summary_ok:
                    logger.warning(
                        "Semantic summary Citex ingestion failed for %s (summary file %s)",
                        filename,
                        summary_filename,
                    )
        elif not settings.CITEX_ENABLED:
            logger.info("Citex ingestion skipped for %s (CITEX_ENABLED=False)", filename)

        # Store file metadata
        file_doc: dict[str, Any] = {
            "file_id": file_id,
            "project_id": project_id,
            "filename": filename,
            "file_type": file_type,
            "content_type": _get_content_type(ext),
            "file_size": len(content),
            "content_hash": content_hash,
            "status": "completed",
            "records_created": records,
            "text_document_id": text_document_id,
            "citex_ingested": result["citex_ingested"],
            "error_message": None,
            "created_at": now,
            "updated_at": now,
        }
        await db.project_files.insert_one(file_doc)

    except Exception as exc:
        logger.exception("Failed to process project file %s: %s", filename, exc)
        result["status"] = "failed"
        result["error_message"] = str(exc)

        # Still store the file metadata with failed status
        file_doc = {
            "file_id": file_id,
            "project_id": project_id,
            "filename": filename,
            "file_type": file_type,
            "content_type": _get_content_type(ext),
            "file_size": len(content),
            "content_hash": content_hash,
            "status": "failed",
            "records_created": 0,
            "text_document_id": None,
            "citex_ingested": False,
            "error_message": str(exc),
            "created_at": now,
            "updated_at": now,
        }
        await db.project_files.insert_one(file_doc)

    return result


async def process_project_note(
    project_id: str,
    title: str,
    content: str,
    db: AsyncIOMotorDatabase,  # type: ignore[type-arg]
) -> dict[str, Any]:
    """
    Process a raw text note pasted from UI and persist semantic insights.
    """
    note_title = title.strip() or "Project Note"
    text = content.strip()
    if not text:
        return {"status": "failed", "error_message": "Note content is empty."}

    document_id = await _store_text_document(
        project_id=project_id,
        source="ui_note",
        source_ref=f"{note_title}:{generate_uuid()[:8]}",
        title=note_title,
        content=text,
        db=db,
    )
    extraction = await extract_semantic_insights(
        project_id=project_id,
        source_type="ui_note",
        source_ref=note_title,
        content=text,
        db=db,
    )
    await generate_project_snapshot(project_id=project_id, db=db, force=True)
    return {
        "status": "completed",
        "document_id": document_id,
        "insights_created": extraction.get("created", 0),
    }


async def retry_project_file_citex_index(
    *,
    project_id: str,
    file_id: str,
    db: AsyncIOMotorDatabase,  # type: ignore[type-arg]
) -> dict[str, Any]:
    """
    Retry Citex indexing for an already-uploaded project file using stored text.

    This path is used when the original ingest attempt failed or was skipped.
    Since raw file bytes are not persisted for all uploads, retry uses the
    canonical ``text_documents`` representation and sends a text-safe filename.
    """
    file_doc = await db.project_files.find_one({"project_id": project_id, "file_id": file_id})
    if not file_doc:
        return {
            "status": "failed",
            "message": f"File '{file_id}' not found in project '{project_id}'.",
            "citex_ingested": False,
            "job_id": None,
        }

    filename = str(file_doc.get("filename") or file_id)
    source_type = str(file_doc.get("file_type") or "documentation")
    text_document_id = str(file_doc.get("text_document_id") or "")

    text_doc = None
    if text_document_id:
        text_doc = await db.text_documents.find_one({"document_id": text_document_id}, {"_id": 0})
    if text_doc is None:
        text_doc = await db.text_documents.find_one(
            {"project_id": project_id, "source": source_type, "source_ref": filename},
            {"_id": 0},
        )

    text_content = str((text_doc or {}).get("content") or "").strip()
    if not text_content:
        await db.project_files.update_one(
            {"project_id": project_id, "file_id": file_id},
            {"$set": {"citex_ingested": False, "error_message": "No extracted text found for retry.", "updated_at": utc_now()}},
        )
        return {
            "status": "failed",
            "message": "No extracted text found; re-upload the file to index again.",
            "citex_ingested": False,
            "job_id": None,
        }

    summary_doc = await db.semantic_summaries.find_one(
        {"project_id": project_id, "source_type": source_type, "source_ref": filename},
        {"summary_text": 1, "_id": 0},
    )
    summary_text = str((summary_doc or {}).get("summary_text") or "").strip()
    citex_content = text_content
    if summary_text:
        citex_content = (
            f"{text_content}\n\n"
            "## ChiefOps Semantic Insights\n"
            f"{summary_text}\n"
        )

    retry_filename = f"{Path(filename).stem}.txt"
    success, job_id = await _retry_ingest_to_citex(
        project_id=project_id,
        content=citex_content,
        source=source_type,
        source_ref=filename,
        retry_filename=retry_filename,
    )

    await db.project_files.update_one(
        {"project_id": project_id, "file_id": file_id},
        {
            "$set": {
                "citex_ingested": success,
                "error_message": None if success else "Citex retry indexing failed.",
                "updated_at": utc_now(),
            }
        },
    )

    return {
        "status": "completed" if success else "failed",
        "message": "Re-indexed in Citex." if success else "Failed to re-index in Citex.",
        "citex_ingested": success,
        "job_id": job_id,
    }


# ---------------------------------------------------------------------------
# File type detection
# ---------------------------------------------------------------------------


def _detect_file_type(ext: str) -> str:
    """Map file extension to a processing category."""
    if ext == ".json":
        return "slack_json"
    if ext == ".xlsx":
        return "jira_xlsx"
    return "documentation"


# ---------------------------------------------------------------------------
# Category-specific processors
# ---------------------------------------------------------------------------


async def _process_slack_json(
    project_id: str,
    filename: str,
    content: bytes,
    db: AsyncIOMotorDatabase,  # type: ignore[type-arg]
) -> tuple[str, int]:
    """Parse a Slack JSON export and build a concatenated text document.

    Returns:
        Tuple of (extracted_text, records_created).
    """
    text = content.decode("utf-8")
    payload = json.loads(text)

    if isinstance(payload, list):
        messages = payload
    elif isinstance(payload, dict):
        candidates = (
            payload.get("messages"),
            payload.get("data"),
            payload.get("items"),
        )
        messages = next((c for c in candidates if isinstance(c, list)), [payload])
    else:
        messages = []

    lines: list[str] = []
    msg_count = 0
    for msg in messages:
        if not isinstance(msg, dict):
            continue
        user = msg.get("user") or msg.get("username") or msg.get("author") or "unknown"
        ts = msg.get("ts") or msg.get("timestamp") or msg.get("time") or ""

        msg_text = str(msg.get("text") or "").strip()
        if not msg_text and isinstance(msg.get("blocks"), list):
            block_texts: list[str] = []
            for block in msg["blocks"]:
                if not isinstance(block, dict):
                    continue
                text_obj = block.get("text")
                if isinstance(text_obj, dict):
                    value = str(text_obj.get("text") or "").strip()
                    if value:
                        block_texts.append(value)
            msg_text = " ".join(block_texts).strip()

        if msg_text:
            lines.append(f"[{ts}] {user}: {msg_text}")
            msg_count += 1

    full_text = "\n".join(lines)
    return full_text, msg_count


async def _process_jira_xlsx(
    project_id: str,
    filename: str,
    content: bytes,
    db: AsyncIOMotorDatabase,  # type: ignore[type-arg]
) -> tuple[str, int]:
    """Parse a Jira XLSX export using openpyxl.

    Reuses column detection and row parsing from jira_csv.py where possible.
    Builds a text document and upserts tasks into jira_tasks.

    Returns:
        Tuple of (extracted_text, records_created).
    """
    from openpyxl import load_workbook

    from app.services.ingestion.jira_csv import (
        _build_column_map,
        _normalise_status,
        _parse_date,
        _parse_row,
        _parse_story_points,
    )

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return "", 0

    # Read all rows as lists of strings
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()

    if len(rows) < 2:
        return "", 0

    # First row is the header
    header = rows[0]
    col_map = _build_column_map(header)

    records_created = 0
    text_lines: list[str] = []
    now = utc_now()

    # _parse_row expects an IngestionFileResult-like object for skip tracking
    from app.services.ingestion.slack_admin import IngestionFileResult

    result = IngestionFileResult(filename=filename, file_type="jira")

    for row_number, row_values in enumerate(rows[1:], start=2):
        if not any(v.strip() for v in row_values):
            continue

        parsed = _parse_row(row_values, col_map, row_number, result)
        if not parsed:
            continue

        task_key = parsed.get("task_key", "")
        if not task_key:
            continue

        # Build task doc following jira_csv pattern
        task_doc: dict[str, Any] = {
            "task_key": task_key,
            "project_key": parsed.get("project_key", task_key.split("-")[0] if "-" in task_key else ""),
            "summary": parsed.get("summary", ""),
            "status": _normalise_status(parsed.get("status", "")),
            "priority": parsed.get("priority", ""),
            "assignee": parsed.get("assignee", ""),
            "reporter": parsed.get("reporter", ""),
            "issue_type": parsed.get("issue_type", ""),
            "story_points": _parse_story_points(parsed.get("story_points", "")),
            "created_date": _parse_date(parsed.get("created", "")),
            "updated_date": _parse_date(parsed.get("updated", "")),
            "due_date": _parse_date(parsed.get("due_date", "")),
            "resolution_date": _parse_date(parsed.get("resolution_date", "")),
            "labels": [l.strip() for l in parsed.get("labels", "").split(",") if l.strip()],
            "components": [c.strip() for c in parsed.get("components", "").split(",") if c.strip()],
            "sprint": parsed.get("sprint", ""),
            "source_project_id": project_id,
            "updated_at": now,
        }

        await db.jira_tasks.update_one(
            {"task_key": task_key},
            {"$set": task_doc, "$setOnInsert": {"created_at": now}},
            upsert=True,
        )
        records_created += 1

        # Build text line for the text document
        text_lines.append(
            f"{task_key}: [{task_doc['status']}] {task_doc['summary']} "
            f"(assignee={task_doc['assignee']}, priority={task_doc['priority']})"
        )

    # Fallback for non-Jira spreadsheets: preserve usable row text for indexing.
    if not text_lines:
        fallback_lines: list[str] = []
        for row in rows[1:]:
            values = [v.strip() for v in row if isinstance(v, str) and v.strip()]
            if not values:
                continue
            fallback_lines.append(" | ".join(values[:12]))
        if fallback_lines:
            return "\n".join(fallback_lines), len(fallback_lines)

    full_text = "\n".join(text_lines)
    return full_text, records_created


async def _process_documentation(
    project_id: str,
    filename: str,
    content: bytes,
    ext: str,
) -> tuple[str, int]:
    """Extract text from a documentation file (PDF, DOCX, MD, TXT).

    Returns:
        Tuple of (extracted_text, records_created). records_created is always
        1 if text was extracted, 0 otherwise.
    """

    class _ErrorCollector:
        """Minimal stand-in for IngestionFileResult to capture extraction errors."""
        errors: list[str] = []

    errors = _ErrorCollector()
    errors.errors = []

    extracted: str | None = None

    if ext in _TEXT_EXTENSIONS:
        extracted = _extract_plain_text(content)
    elif ext == ".pdf":
        extracted = _extract_pdf_text(content, errors)  # type: ignore[arg-type]
    elif ext == ".docx":
        # python-docx needs a file path, write to temp
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            extracted = _extract_docx_text(tmp_path, errors)  # type: ignore[arg-type]
        finally:
            Path(tmp_path).unlink(missing_ok=True)
    else:
        # Fallback: try plain text decoding
        extracted = _extract_plain_text(content)

    if errors.errors:
        logger.warning("Extraction warnings for %s: %s", filename, errors.errors)

    if extracted and extracted.strip():
        return extracted, 1
    return "", 0


# ---------------------------------------------------------------------------
# Text document storage
# ---------------------------------------------------------------------------


async def _store_text_document(
    project_id: str,
    source: str,
    source_ref: str,
    title: str,
    content: str,
    db: AsyncIOMotorDatabase,  # type: ignore[type-arg]
) -> str:
    """Upsert a text_documents entry tagged with project_id.

    Returns:
        The document_id of the created/updated text document.
    """
    content_hash = compute_hash(content.encode("utf-8"))
    document_id = generate_uuid()
    now = utc_now()

    doc: dict[str, Any] = {
        "document_id": document_id,
        "project_id": project_id,
        "source": source,
        "source_ref": source_ref,
        "title": title,
        "content": content,
        "content_hash": content_hash,
        "content_type": "text/plain",
        "created_at": now,
        "updated_at": now,
    }

    # Upsert by project_id + source + source_ref to avoid duplicates on re-upload
    result = await db.text_documents.update_one(
        {"project_id": project_id, "source": source, "source_ref": source_ref},
        {"$set": doc},
        upsert=True,
    )

    # If it was an update, retrieve the existing document_id
    if not result.upserted_id:
        existing = await db.text_documents.find_one(
            {"project_id": project_id, "source": source, "source_ref": source_ref},
            {"document_id": 1},
        )
        if existing:
            document_id = existing["document_id"]

    return document_id


# ---------------------------------------------------------------------------
# Citex ingestion
# ---------------------------------------------------------------------------


async def _ingest_to_citex(
    project_id: str,
    content: str,
    filename: str,
    source: str,
    *,
    source_ref: str | None = None,
    file_bytes: bytes | None = None,
    content_type: str | None = None,
) -> bool:
    """Attempt to ingest content into Citex. Returns True on success."""
    settings = get_settings()
    citex_project_id = derive_citex_project_id(
        configured_project_id=settings.CITEX_PROJECT_ID,
        api_key=settings.CITEX_API_KEY,
        fallback_project_id=project_id,
    )
    client = CitexClient(
        settings.CITEX_API_URL,
        api_key=settings.CITEX_API_KEY,
        user_id=settings.CITEX_USER_ID,
        scope_id=f"project:{project_id}",
    )

    try:
        if not await client.ping():
            logger.debug("Citex unavailable; skipping ingestion for %s", filename)
            return False

        metadata = {
            "source": source,
            "source_ref": source_ref or filename,
            "upload_type": "project_file",
        }
        response = await client.ingest_document(
            project_id=citex_project_id,
            content=content,
            metadata=metadata,
            filename=filename,
            file_bytes=file_bytes,
            content_type=content_type,
        )
        return bool(response)

    except Exception as exc:
        logger.warning("Citex ingestion failed for %s: %s", filename, exc)
        return False
    finally:
        await client.close()


async def _record_citex_ingestion_state(
    *,
    db: AsyncIOMotorDatabase,  # type: ignore[type-arg]
    project_id: str,
    source_group: str,
    source: str,
    source_ref: str,
    document_id: str,
    content_hash: str,
) -> None:
    """Persist Citex ingestion state used by project analysis dedupe."""
    settings = get_settings()
    citex_project_id = derive_citex_project_id(
        configured_project_id=settings.CITEX_PROJECT_ID,
        api_key=settings.CITEX_API_KEY,
        fallback_project_id=project_id,
    )
    now = utc_now()
    await db.citex_ingestion_state.update_one(
        {
            "project_id": project_id,
            "source_group": source_group,
            "document_id": document_id,
        },
        {
            "$set": {
                "project_id": project_id,
                "citex_project_id": citex_project_id,
                "source_group": source_group,
                "source": source,
                "source_ref": source_ref,
                "document_id": document_id,
                "content_hash": content_hash,
                "status": "ingested",
                "last_ingested_at": now,
                "updated_at": now,
            },
            "$setOnInsert": {"created_at": now},
        },
        upsert=True,
    )


async def _retry_ingest_to_citex(
    *,
    project_id: str,
    content: str,
    source: str,
    source_ref: str,
    retry_filename: str,
) -> tuple[bool, str | None]:
    """Retry Citex indexing from text content and return (success, job_id)."""
    settings = get_settings()
    citex_project_id = derive_citex_project_id(
        configured_project_id=settings.CITEX_PROJECT_ID,
        api_key=settings.CITEX_API_KEY,
        fallback_project_id=project_id,
    )
    client = CitexClient(
        settings.CITEX_API_URL,
        api_key=settings.CITEX_API_KEY,
        user_id=settings.CITEX_USER_ID,
        scope_id=f"project:{project_id}",
    )
    try:
        if not await client.ping():
            return False, None

        metadata = {
            "source": source,
            "source_ref": source_ref,
            "upload_type": "project_file_retry",
            "original_filename": source_ref,
        }
        response = await client.ingest_document(
            project_id=citex_project_id,
            content=content,
            metadata=metadata,
            filename=retry_filename,
            content_type="text/plain",
        )
        if not response:
            return False, None
        return True, str(response.get("jobId")) if response.get("jobId") else None
    except Exception as exc:
        logger.warning("Citex retry ingestion failed for %s: %s", source_ref, exc)
        return False, None
    finally:
        await client.close()
